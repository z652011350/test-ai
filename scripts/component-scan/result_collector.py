"""
result_collector.py - 部件扫描结果收集与汇总

扫描输出目录，收集各部件的扫描结果，生成 JSONL/XLSX/Markdown 汇总报告。
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


def scan_component_results(out_base: Path) -> List[Dict]:
    """
    扫描输出目录，收集各部件的扫描状态。

    Args:
        out_base: 输出根目录

    Returns:
        结果列表，每项包含 component_name, status, report_files 等
    """
    results = []
    if not out_base.exists():
        return results

    for component_dir in sorted(out_base.iterdir()):
        if not component_dir.is_dir():
            continue

        component_name = component_dir.name
        result = {
            "component_name": component_name,
            "kit_name": "",
            "status": "unknown",
            "report_files": [],
            "scan_time": "",
            "duration_seconds": 0,
            "analyze_depth": "",
        }

        # 检查报告文件
        report_patterns = [
            "*_ERROR_CODE_ANALYSIS_REPORT.md",
            "*_API_CALL_CHAIN_AND_EXCEPTION_REPORT.md",
            "*_DOC_CONSISTENCY_REPORT.md",
            "*_ERROR_CODE_MESSAGE.md",
            "*_ISSUE_REPORT.md",
        ]

        found_reports = []
        for pattern in report_patterns:
            matches = list(component_dir.glob(pattern))
            found_reports.extend([str(m) for m in matches])

        result["report_files"] = found_reports

        # 判断状态
        issue_reports = [f for f in found_reports if "ISSUE_REPORT" in f]
        if issue_reports:
            result["status"] = "success"
        elif found_reports:
            result["status"] = "partial"
        else:
            result["status"] = "no_reports"

        # 读取扫描元数据（如果存在）
        meta_file = component_dir / "_scan_meta.json"
        if meta_file.exists():
            try:
                with open(meta_file, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                result["kit_name"] = meta.get("kit_name", "")
                result["scan_time"] = meta.get("scan_time", "")
                result["duration_seconds"] = meta.get("duration_seconds", 0)
                result["analyze_depth"] = meta.get("analyze_depth", "")
                if meta.get("status") == "failed":
                    result["status"] = "failed"
            except (json.JSONDecodeError, IOError):
                pass

        results.append(result)

    return results


def generate_jsonl(results: List[Dict], output_path: Path) -> None:
    """生成 JSONL 汇总文件。"""
    with open(output_path, "w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"[result_collector] JSONL 汇总: {output_path} ({len(results)} 条)")


def generate_xlsx(results: List[Dict], output_path: Path) -> None:
    """生成 XLSX 汇总表格。"""
    try:
        import openpyxl
    except ImportError:
        print("[result_collector] 跳过 XLSX 生成: 未安装 openpyxl (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "扫描汇总"

    # 表头
    headers = ["部件名称", "Kit", "状态", "分析深度", "报告数量", "耗时(秒)", "扫描时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # 数据行
    status_map = {
        "success": "成功",
        "partial": "部分完成",
        "failed": "失败",
        "no_reports": "无报告",
        "skipped": "已跳过",
        "unknown": "未知",
    }

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["component_name"])
        ws.cell(row=row_idx, column=2, value=r.get("kit_name", ""))
        ws.cell(row=row_idx, column=3, value=status_map.get(r["status"], r["status"]))
        ws.cell(row=row_idx, column=4, value=r.get("analyze_depth", ""))
        ws.cell(row=row_idx, column=5, value=len(r.get("report_files", [])))
        ws.cell(row=row_idx, column=6, value=r.get("duration_seconds", 0))
        ws.cell(row=row_idx, column=7, value=r.get("scan_time", ""))

    # 调整列宽
    col_widths = [35, 20, 12, 12, 10, 10, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(output_path)
    print(f"[result_collector] XLSX 汇总: {output_path} ({len(results)} 条)")


def generate_markdown(results: List[Dict], output_path: Path) -> None:
    """生成 Markdown 汇总报告。"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    status_emoji = {
        "success": "OK",
        "partial": "partial",
        "failed": "FAIL",
        "no_reports": "-",
        "skipped": "SKIP",
        "unknown": "?",
    }

    lines = [
        f"# 部件扫描汇总报告",
        f"",
        f"生成时间: {now}",
        f"",
        f"## 统计概览",
        f"",
    ]

    total = len(results)
    success = sum(1 for r in results if r["status"] == "success")
    failed = sum(1 for r in results if r["status"] == "failed")
    partial = sum(1 for r in results if r["status"] == "partial")
    no_reports = sum(1 for r in results if r["status"] == "no_reports")
    skipped = sum(1 for r in results if r["status"] == "skipped")

    lines.extend([
        f"| 指标 | 数量 |",
        f"|------|------|",
        f"| 总部件数 | {total} |",
        f"| 扫描成功 | {success} |",
        f"| 部分完成 | {partial} |",
        f"| 扫描失败 | {failed} |",
        f"| 无报告 | {no_reports} |",
        f"| 已跳过 | {skipped} |",
        f"",
        f"## 各部件扫描详情",
        f"",
        f"| 部件名称 | Kit | 状态 | 分析深度 | 报告数 | 耗时(s) | 扫描时间 |",
        f"|----------|-----|------|----------|--------|---------|----------|",
    ])

    for r in results:
        status_str = status_emoji.get(r["status"], r["status"])
        kit = r.get("kit_name", "")
        depth = r.get("analyze_depth", "")
        report_count = len(r.get("report_files", []))
        duration = r.get("duration_seconds", 0)
        scan_time = r.get("scan_time", "")
        if isinstance(duration, (int, float)):
            duration = f"{duration:.0f}"

        lines.append(
            f"| {r['component_name']} | {kit} | {status_str} | {depth} | "
            f"{report_count} | {duration} | {scan_time} |"
        )

    # 失败部件详情
    failed_results = [r for r in results if r["status"] == "failed"]
    if failed_results:
        lines.extend(["", "## 失败部件", ""])
        for r in failed_results:
            lines.append(f"- **{r['component_name']}**")

    lines.append("")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[result_collector] Markdown 汇总: {output_path}")


def collect_and_summarize(out_base: Path) -> None:
    """主入口：收集结果并生成所有汇总文件。"""
    print(f"\n{'=' * 60}")
    print("[result_collector] 开始收集扫描结果...")

    results = scan_component_results(out_base)

    if not results:
        print("[result_collector] 未找到任何扫描结果")
        return

    out_base.mkdir(parents=True, exist_ok=True)
    generate_jsonl(results, out_base / "scan_summary.jsonl")
    generate_xlsx(results, out_base / "scan_summary.xlsx")
    generate_markdown(results, out_base / "scan_summary.md")

    print(f"[result_collector] 汇总完成，共 {len(results)} 个部件")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="部件扫描结果收集与汇总")
    parser.add_argument("-out_base", required=True, help="输出根目录")
    args = parser.parse_args()
    collect_and_summarize(Path(args.out_base))
